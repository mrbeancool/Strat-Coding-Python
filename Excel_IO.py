'''
This file has Excel_IO class and related functions
'''
import datetime
import logging
import configparser # parameters.ini
import openpyxl # Excel file
import csv

class Excel_IO():
    config = configparser.ConfigParser()
    filepath = None
    filename = None
    wb = None
    last_oid = None
    # sheetname = None

    def __init__(self):
        self.config.read('parameters.ini')

        loglevel = self.config['LOG']['level']
        logpath = self.config['LOG']['path']
        logpath += '.' + str(datetime.date.today()) + '.log'

        # DEBUG, INFO, WARNING, ERROR, CRITICAL
        logging.basicConfig(filename=logpath, encoding='utf-8', level=loglevel,
                            format='%(asctime)s %(levelname)s: %(message)s')
        self.filepath = self.config['EXCEL']['filepath']
        self.filename = self.config['EXCEL']['filename']
        self.last_oid = self.config['TWS']['lastorderid']

        # Load the Excel workbook
        self.wb = openpyxl.load_workbook(self.filepath+self.filename)

    def Setup(self, choice):
    # open Excel sheet based on user choice
        match choice:
            case 1:
                sheetname =  self.config['EXCEL']['existingorderssheet']
            case 2:
                sheetname = self.config['EXCEL']['metasheet']
            case 3:
                sheetname = self.config['EXCEL']['placeorderssheet']
            case 4:
                sheetname =  self.config['EXCEL']['placeorderssheet']
            case 5:
                sheetname = self.config['EXCEL']['metasheet']

        logging.info(f"Opened Excel file: {self.filepath} {self.filename} sheet: {sheetname}")

        # Select the specific sheet
        sheet = self.wb[sheetname]
        logging.info("Setup complete")
        return sheet

    def GetExistingOrders(self, list, sheet):
        # delete data from the sheet except header row
        sheet.delete_rows(2, sheet.max_row)
        # load from list to sheet
        for row_idx, row_data in enumerate(list, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        self.wb.save(self.filepath+self.filename)
        return

    def TransferExcelOrders(self, sheet, orderStatus):
    # transfer orders from sheet to list
        list1 = [] # Reset list
        lr = 0  # to store no of rows
        firstRow = True

        for row in sheet.iter_rows(values_only=True):
            if (firstRow):  # Ignore the header row
                firstRow = False
                continue
            if(row[8] != orderStatus): # ignore order where status is not New
                continue
            list1.append(list(row))
            lr += 1

        logging.info(f"TransferExcelOrders successful, {lr} rows transferred to list")
        list1 = LogicOps.TruncateList(list1, 12)
        return list1

    def SetLastOrderId(self, lastorderid):
        self.config = configparser.ConfigParser()
        self.config.read('parameters.ini')
        self.config.set('TWS','lastorderid',str(lastorderid))
        with open('parameters.ini','w') as configfile:
            self.config.write(configfile)

    def GetTickerList (self, sheet):
        # transfer tickers from sheet to list
        list1 = []  # Reset list
        lr = 0  # to store no of rows
        firstRow = True

        for row in sheet.iter_rows(values_only=True):
            if (firstRow):  # Ignore the header row
                firstRow = False
                continue
            list1.append(list(row))
            lr += 1
        logging.info(f"GetTickerList successful, {lr} rows transferred to list")
        return list1

    def GetStratPatterns(self, sheet):
        list1 = []
        firstRow = True

        self.config.read('parameters.ini')
        mt4filepath = self.config['MT4']['filepath']

        # Get the data from MT4 file to a list
        with open(mt4filepath, 'r', newline='') as csv_file:
            csv_reader = csv.reader(csv_file)
            for row in csv_reader:
                list1.append(row)

        # Run loop on Excel and find the data in list
        # todo why enumerate runs on list1
        # shouldnt it run on sheet?
        for row_idx, row_data in enumerate(list1, start=2):
            excelCurr = sheet.cell(row_idx, 1).value
            for lr in list1:
                # 1st item in the row has the symbol
                if (lr[0] == excelCurr):
                    sheet.cell(row=row_idx, column=5, value=lr[1])
                    sheet.cell(row=row_idx, column=6, value=lr[2])
                    sheet.cell(row=row_idx, column=7, value=lr[3])
                    sheet.cell(row=row_idx, column=8, value=lr[4])
                    sheet.cell(row=row_idx, column=9, value=lr[5])
                    break
        self.wb.save(self.filepath+self.filename)
        return
class LogicOps():
    @classmethod
    def RemoveDups(self, origlist):
        nodup = []
        for row in origlist:
            if row not in nodup:
                nodup.append(row)
        return nodup

    @classmethod
    def TruncateList(self, origlist, n):
        for row in origlist:
            row[:] = row[:n]
        return origlist